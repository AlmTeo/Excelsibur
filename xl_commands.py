'''
from openpyxl import Workbook
xlsx = Workbook()

# grab the active worksheet
worksheet = xlsx.active

# Data can be assigned directly to cells
worksheet['A1'] = 42

# Rows can also be appended
worksheet.append([1, 2, 3])

# Python types will automatically be converted
import datetime
worksheet['A2'] = datetime.datetime.now()

# Save the file
xlsx.save("sample.xlsx")
'''



###########################################################

# Opening Excel Documents
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
type(wb) 


# Getting Sheets from the Workbook
wb = openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())


sheet = wb.get_sheet_by_name('Sheet')
print(sheet)

print(type(sheet))
print(sheet.title)


anotherSheet = wb.active
print(anotherSheet)



# ##################################################


# print(sheet['A1'].value)
# print(sheet['B1'].value)
# print(sheet['C1'].value)


# Specific cell data
# c = sheet['B3']
# print('Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value)


# All rows in range
# for rowOfCellObjects in sheet['A2':'C5']:
# 	for cellObj in rowOfCellObjects:
# 		print(cellObj.coordinate, cellObj.value)
# 	print('--- END OF ROW ---')